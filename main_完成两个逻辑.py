
from SqlHelper import MysqlHelp
import openpyxl

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.

temp_pocess = 'HS'
if __name__ == '__main__':
    sql = " EXEC SMT1.SMT1.SP_DataInspectAoiSpi "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)
    #qdata = [('BCLW35', '变形', '1'), ('BCLW35', '缺件', '2'), ('BCLW35', '漏焊', '3'), ('BCLW35', '极性反', '4'),
    #         ('BEAF05', '元件破损', '5'), ('BEAF05', '缺件', '6'), ('BEAF09', '元件破损', '7'), ('BEAF09', '缺件', '8')]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 15
    sheet['A1']='Compare AOI and SPI Data'
    sheet['A2']='AOI表中存在，但是SPI表中不存在'
    print(len(qdata))
    sheet.append(['STARTTIME','BARCODE','WO','ItemCode','LINE','SIDE'])

    for i in range(0,len(qdata)):
        sheet.append(qdata[i])

    sheet2 = workbook.create_sheet('sheet2')
    sheet2.column_dimensions["A"].width = 15
    sheet2.column_dimensions["B"].width = 20
    sheet2.column_dimensions["C"].width = 15
    sheet2['A1']='Compare AOI and SPI Data'
    sheet2['A2']='SPI表中存在，但是AOI表中经过15天仍不存在'
    print(len(qdata))
    sheet2.append(['STARTTIME','BARCODE','WO','ItemCode','LINE','SIDE'])

    sql = " EXEC SMT1.SMT1.SP_DataInspectSpiAoi "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)

    for i in range(0,len(qdata)):
        sheet2.append(qdata[i])

    sql = " EXEC SMT1.SMT1.SP_DataInspectSpiSideChk "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)


    sheet3 = workbook.create_sheet('sheet3')
    sheet3.column_dimensions["A"].width = 15
    sheet3.column_dimensions["B"].width = 20
    sheet3.column_dimensions["C"].width = 15
    #sheet3['A1']='Compare AOI and SPI Data'
    sheet3['A2']='检查SPI数据记录中，两面数据是否都存在'
    print(len(qdata))
    sheet3.append(['STARTTIME','BARCODE','WO','ITEM_CODE','LINE','SIDE'])

    for i in range(0,len(qdata)):
        sheet3.append(qdata[i])



    sql = " EXEC SMT1.SMT1.SP_DataInspectAoiSideChk "
    #sql = " EXEC SMT1.testproc "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)


    sheet4 = workbook.create_sheet('sheet4')
    sheet4.column_dimensions["A"].width = 15
    sheet4.column_dimensions["B"].width = 20
    sheet4.column_dimensions["C"].width = 15
    #sheet4['A1']='Compare AOI and SPI Data'
    sheet4['A2']='检查AOI数据记录中，两面数据是否都存在'
    print(len(qdata))
    sheet4.append(['STARTTIME','BARCODE','WO','ITEM_CODE','LINE','SIDE'])

    for i in range(0,len(qdata)):
        sheet4.append(qdata[i])

    sheet.title='仅在aoi中存在'
    sheet2.title = '仅在spi中存在'
    sheet3.title = 'spi中单面数据'
    sheet4.title = 'aoi中单面数据'

    workbook.save('DataInspectionReport.xlsx')
