
from SqlHelper import MysqlHelp
import openpyxl
from openpyxl import load_workbook

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.

temp_pocess = 'HS'
if __name__ == '__main__':
    sql = " EXEC SMT1.SMT1.SP_DataInspectAoiSpi "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)
    #qdata = [('BCLW35', '变形', '1'), ('BCLW35', '缺件', '2'), ('BCLW35', '漏焊', '3'), ('BCLW35', '极性反', '4'),
    #         ('BEAF05', '元件破损', '5'), ('BEAF05', '缺件', '6'), ('BEAF09', '元件破损', '7'), ('BEAF09', '缺件', '8')]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 15
    sheet['A1']='Compare AOI and SPI Data'
    sheet['A2']='AOI表中存在，但是SPI表中不存在'
    print(len(qdata))
    sheet.append(['STARTTIME','BARCODE','WO','ItemCode','LINE','SIDE'])

    for i in range(0,len(qdata)):
        sheet.append(qdata[i])

    sheet2 = workbook.create_sheet('sheet2')
    sheet2.column_dimensions["A"].width = 15
    sheet2.column_dimensions["B"].width = 20
    sheet2.column_dimensions["C"].width = 15
    sheet2['A1']='Compare AOI and SPI Data'
    sheet2['A2']='SPI表中存在，但是AOI表中经过15天仍不存在'
    print(len(qdata))
    sheet2.append(['STARTTIME','BARCODE','WO','ItemCode','LINE','SIDE'])

    sql = " EXEC SMT1.SMT1.SP_DataInspectSpiAoi "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)

    for i in range(0,len(qdata)):
        sheet2.append(qdata[i])

    sql = " EXEC SMT1.SMT1.SP_DataInspectSpiSideChk "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)


    sheet3 = workbook.create_sheet('sheet3')
    sheet3.column_dimensions["A"].width = 15
    sheet3.column_dimensions["B"].width = 20
    sheet3.column_dimensions["C"].width = 15
    #sheet3['A1']='Compare AOI and SPI Data'
    sheet3['A2']='检查SPI数据记录中，两面数据是否都存在'
    print(len(qdata))
    sheet3.append(['STARTTIME','BARCODE','WO','ITEM_CODE','LINE','SIDE'])

    for i in range(0,len(qdata)):
        sheet3.append(qdata[i])


    ####################################################
    ####################################################
    sql = " EXEC SMT1.SMT1.SP_DataInspectAoiSideChk "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)


    sheet4 = workbook.create_sheet('sheet4')
    sheet4.column_dimensions["A"].width = 15
    sheet4.column_dimensions["B"].width = 20
    sheet4.column_dimensions["C"].width = 15
    #sheet4['A1']='Compare AOI and SPI Data'
    sheet4['A2']='检查AOI数据记录中，两面数据是否都存在'
    print(len(qdata))
    sheet4.append(['STARTTIME','BARCODE','WO','ITEM_CODE','LINE','SIDE'])

    for i in range(0,len(qdata)):
        sheet4.append(qdata[i])


    ####################################################
    #L1每个工单上工时列表
    ####################################################
    sql = " EXEC SMT1.SMT1.SP_WorkTimeL1 "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)

    sheet5 = workbook.create_sheet('sheet5')
    sheet5.column_dimensions["A"].width = 15
    sheet5.column_dimensions["B"].width = 20
    sheet5.column_dimensions["C"].width = 25
    sheet5.column_dimensions["D"].width = 25
    #sheet5['A1']='Compare AOI and SPI Data'
    print(len(qdata))
    sheet5.append(['WO','BARCODE','AOI结束时间','SPI开始时间','总时间 /秒'])

    for i in range(0,len(qdata)):
        sheet5.append(qdata[i])



    ####################################################
    #L2每个工单上工时列表
    ####################################################
    sql = " EXEC SMT1.SMT1.SP_WorkTimeL2 "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)

    sheet6 = workbook.create_sheet('sheet6')
    sheet6.column_dimensions["A"].width = 15
    sheet6.column_dimensions["B"].width = 20
    sheet6.column_dimensions["C"].width = 25
    sheet6.column_dimensions["D"].width = 25
    #sheet6['A1']='Compare AOI and SPI Data'
    print(len(qdata))
    sheet6.append(['WO','BARCODE','AOI结束时间','SPI开始时间','总时间 /秒'])

    for i in range(0,len(qdata)):
        sheet6.append(qdata[i])


    ####################################################
    #L1每个工单上工时统计
    ####################################################
    sql = " EXEC SMT1.SMT1.SP_WorkTimeL1Total "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)

    sheet7 = workbook.create_sheet('sheet7')
    sheet7.column_dimensions["A"].width = 15
    sheet7.column_dimensions["B"].width = 15
    sheet7.column_dimensions["C"].width = 25
    sheet7.column_dimensions["D"].width = 25
    #sheet7['A1']='Compare AOI and SPI Data'
    #sheet7['A2']='检查AOI数据记录中，两面数据是否都存在'
    print(len(qdata))
    sheet7.append(['WO','总工时 /秒'])

    for i in range(0,len(qdata)):
        sheet7.append(qdata[i])


    ####################################################
    #L2每个工单上工时统计
    ####################################################
    sql = " EXEC SMT1.SMT1.SP_WorkTimeL2Total "
    qdata = MysqlHelp().select_exec_sp(sql)
    print("###############")
    print(sql)
    print(qdata)

    sheet8 = workbook.create_sheet('sheet8')
    sheet8.column_dimensions["A"].width = 15
    sheet8.column_dimensions["B"].width = 15
    sheet8.column_dimensions["C"].width = 25
    sheet8.column_dimensions["D"].width = 25
    #sheet8['A1']='Compare AOI and SPI Data'
    #sheet8['A2']='检查AOI数据记录中，两面数据是否都存在'
    print(len(qdata))
    sheet8.append(['WO','总工时 /秒'])

    for i in range(0,len(qdata)):
        sheet8.append(qdata[i])

    sheet.title='仅在aoi中存在'
    sheet2.title = '仅在spi中存在'
    sheet3.title = 'spi中单面数据'
    sheet4.title = 'aoi中单面数据'
    sheet5.title = 'LINE1 工时列表'
    sheet6.title = 'LINE2 工时列表'
    sheet7.title = 'LINE1工时统计'
    sheet8.title = 'LINE2工时统计'

    #wb = openpyxl.load_workbook('DataInspectionReport.xlsx')

    sheet9 = workbook.create_sheet('sheet9')
    sheet9.append(['WO', '总工时 /秒'])
    sheet9.column_dimensions["A"].width = 20

    # 选择要处理的 sheet（默认为第一个）
    ws1 = workbook['LINE1工时统计']
    ws2 = workbook['LINE2工时统计']
    ws3 = workbook['sheet9']
    i = 0
    j = 0
    k = 1
    a = ws1.max_row
    b = ws2.max_row
    c = ws3.max_row

    #将ws1，ws2中相同的WO将时间相加，结果放到sheet9中
    for i in range(1, a):
        for j in range(1, b+1):
            if ws1.cell(row=i + 1, column=1).value == ws2.cell(row=j + 1, column=1).value:
                #copy WO
                ws3.cell(row=k + 1, column=1).value = ws1.cell(row=i + 1, column=1).value
                #copy 时间
                ws3.cell(row=k + 1, column=2).value = int(ws1.cell(row=i + 1, column=2).value) + int(
                    ws2.cell(row=j + 1, column=2).value)
                k = k + 1

    c = ws3.max_row
    #将ws1中独有的WO的时间放到sheet9中
    for i in range(1, a):
        for j in range(1,c+1):
            if ws1.cell(row=i + 1, column=1).value == ws3.cell(row=j + 1, column=1).value and ws1.cell(row=i + 1, column=1).value !=None:
                break
        if j==c:
            # copy WO
            ws3.cell(row=k + 1, column=1).value = ws1.cell(row=i + 1, column=1).value
            # copy 时间
            if ws1.cell(row=i+1,column=2).value!=None:
                ws3.cell(row=k+1, column=2).value = int(ws1.cell(row=i+1,column=2).value)
            k = k + 1


    #将ws2中独有的WO的时间放到sheet9中
    #c = ws3.max_row
    for i in range(1, a):
        for j in range(1,c+1):
            if ws2.cell(row=i + 1, column=1).value == ws3.cell(row=j + 1, column=1).value and ws2.cell(row=i + 1, column=1).value !=None:
                break
        if j==c:
            # copy WO
            ws3.cell(row=k + 1, column=1).value = ws2.cell(row=i + 1, column=1).value
            # copy 时间
            #print(ws2.cell(row=i+1,column=2).value)
            if ws2.cell(row=i+1,column=2).value!=None:
                ws3.cell(row=k+1, column=2).value = int(ws2.cell(row=i+1,column=2).value)
            k = k + 1


    sheet9.title = 'WO工时统计'

    workbook.save('DataInspectionReport.xlsx')


